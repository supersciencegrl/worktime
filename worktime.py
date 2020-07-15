from datetime import datetime, timedelta
import statistics as stats

import openpyxl
from winevt import EventLog

def printerrordates():
    for d in lod:
        if len(d['wakelist']) != len(d['sleeplist']):
            print(d['date'])
            print('wake:', d['wakelist'])
            print('sleep:', d['sleeplist'], '\n')

def timestring(time):
    if isinstance(time, datetime):
        value = datetime.strftime(time, '%H:%M:%S')
    else:
        value = 0
    return value

def datestring(date):
    value = datetime.strftime(date, '%a %d-%m-%Y')
    return value

def totime(time):
    value = datetime.strptime(time, '%H:%M:%S.%fZ')
    return value

def ss_tohhmm(seconds):
    hours = int(seconds / 60 // 60)
    mins = int(round(seconds / 60 - (hours * 60)))
    return hours, mins

query = EventLog.Query('System', 'Event/System')

sleepy_tiems = []
waek_tiems = []

for event in query:
    try:
        if event.System.EventID.cdata == '7025':
            newtime = event.System.TimeCreated['SystemTime'] # Format: 2020-07-01T22:13:45.346031400Z
            waek_tiems.append(newtime[:23] + newtime[29:]) # Format: '2020-07-01T22:13:45.346Z'
        elif event.System.EventID.cdata == '42':
            newtime = event.System.TimeCreated['SystemTime'] # Format: 2020-07-01T22:13:45.346031400Z
            sleepy_tiems.append(newtime[:23] + newtime[29:]) # Format: '2020-07-01T22:13:45.346Z'
    except AttributeError:
        pass

print(f'len(sleepy_tiems): {len(sleepy_tiems)}')
print(f'len(waek_tiems): {len(waek_tiems)}\n')

currentevtdate = waek_tiems[0][:10]
sleeptimeindex = 0
waketimeindex = 0

alldates = set([tiem[:10] for tiem in waek_tiems + sleepy_tiems])
alldates = sorted(list(alldates))
lod = []

for date in alldates:
    datedict = {'date': datetime.strptime(date, '%Y-%m-%d'), 'wakelist': [], 'sleeplist': []}
    for n, w in enumerate(waek_tiems[waketimeindex:]):
        if w.startswith(date):
            datedict['wakelist'].append(w[11:])
        else:
            waketimeindex += n
            break
    for n, s in enumerate(sleepy_tiems[sleeptimeindex:]):
        if s.startswith(date):
            datedict['sleeplist'].append(s[11:])
        else:
            sleeptimeindex += n
            break
    lod.append(datedict)

for d in lod:
    d['totaltime'] = 0
    d['dayofweek'] = d['date'].weekday() # Returns integer: Monday == 0
    thisdate = datestring(d['date'])
    firstwtime, firststime = 0, 0
    if d['wakelist']:
        firstwtime = datetime.strptime(d['wakelist'][0], '%H:%M:%S.%fZ')
    if d['sleeplist']:
        firststime = datetime.strptime(d['sleeplist'][0], '%H:%M:%S.%fZ')
    if len(d['wakelist']) == len(d['sleeplist']) and not (firststime - firstwtime).days: # if firststime is after firstwtime: else days = -1
        for n, wtime in enumerate(d['wakelist']):
            if not (totime(d['sleeplist'][n]) - totime(wtime)).days:
                d['totaltime'] += (totime(d['sleeplist'][n]) - totime(wtime)).seconds
    elif len(d['sleeplist']) - len(d['wakelist']) == 1 and (firststime - firstwtime).days == -1: # if firststime is between midnight and waking
        for n, wtime in enumerate(d['wakelist']):
            d['totaltime'] += (totime(d['sleeplist'][n+1]) - totime(wtime)).seconds # remove firststime from effective sleeplist
    elif len(d['wakelist']) - len(d['sleeplist']) == 1 and d['sleeplist']:
        lastwtime = datetime.strptime(d['wakelist'][-1], '%H:%M:%S.%fZ')
        laststime = datetime.strptime(d['sleeplist'][-1], '%H:%M:%S.%fZ')
        if (lastwtime - laststime).days == 0:
            for n, stime in enumerate(d['sleeplist']):
                d['totaltime'] += (totime(stime) - totime(d['wakelist'][n])).seconds # remove lastwtime from effective wakelist

    if d['totaltime']:
        hours, mins = ss_tohhmm(d['totaltime'])
        print(f'{thisdate}\t{hours} h {mins:02d} mins')

firstdate = lod[0]['date']
alldtdates = [d['date'] for d in lod]

averageweekday = [i for i in [d['totaltime'] for d in lod if d['dayofweek'] < 5] if i > 0]
hours, mins = ss_tohhmm(stats.mean(averageweekday))

print(f'\nAverage time per weekday:\t{hours} h {mins:02d} mins\n')

# Load timesheet
inputfile = 'C:\\Users\\S1024501\\OneDrive - Syngenta\\Documents\\Documents\\Timesheet.xlsx'
wb = openpyxl.load_workbook(filename = inputfile, data_only = True)
ws = wb['Full']

weeklist = []
max_row = ws.max_row
firstworkday = datetime(2020, 5, 25, 0, 0)

mydate = firstworkday
todaydateonly = datetime.today().replace(hour = 0).replace(minute = 0).replace(second = 0).replace(microsecond = 0)

for row in range(2, max_row + 1):
    weeklist.append(ws.cell(row = row, column = 1).value)

while mydate != todaydateonly:
    if mydate == firstworkday:
        weeklycomputertime = 0
        lastweek = mydate
    elif not mydate.weekday(): # If it's Monday but not my first day
        hours, mins = ss_tohhmm(weeklycomputertime)
        print(f'Week commencing: {datestring(lastweek)}\t\tComputer time: {hours} h {mins} mins')
        lastweek = mydate
        weeklycomputertime = 0
    if mydate in alldtdates:
        weeklycomputertime += lod[alldtdates.index(mydate)]['totaltime']
    mydate = mydate + timedelta(days = 1)
    
